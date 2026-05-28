"""
patch_fees_from_xlsx.py — overlay per-ticker commission fees from the US broker
transaction xlsx onto data/holdings_cost.json.

Background:
  parse_broker_pdf.py derives per-position `fees` from the PDF "P&L breakdown"
  page. That page sometimes misattributes fees for re-opened tickers (e.g. RKLB
  ended up with $0 fees on the open lot even though 7 RKLB trades each cost $5).
  The xlsx Bookings sheet has every commission line by Trade ID, and the Trades
  sheet labels each Trade ID as "To Open" or "To Close", so we can split fees
  precisely between the still-open lot and the closed tranche.

Usage:
    python3 scripts/patch_fees_from_xlsx.py <path-to-transactions.xlsx>
        [--cost-file data/holdings_cost.json]
        [--dry-run]

What it does:
  • Reads xlsx Bookings + Trades.
  • Builds per-ticker {open_comm, close_comm} aggregates.
  • For every us.open[].tk present in xlsx → sets fees = open_comm.
  • For every us.closed[].tk present in xlsx → sets _costs_paid = close_comm
    (realised stays untouched; PDF realised is already net of fees).
  • Tickers absent from xlsx (older closed lots etc.) are left untouched.
  • Total commission delta is printed for sanity.

The xlsx never leaves your machine — only holdings_cost.json is committed.
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent


def split_commissions(xl_path: Path) -> tuple[dict[str, float], dict[str, float]]:
    """Returns (per_ticker_open_comm, per_ticker_close_comm) from xlsx."""
    wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
    if "Trades" not in wb.sheetnames or "Bookings" not in wb.sheetnames:
        raise SystemExit(
            f"xlsx missing required sheets — found {wb.sheetnames}, "
            "expected at least Trades + Bookings"
        )

    # ── Trades: trade_id → {direction, symbol_base} ──
    ws_t = wb["Trades"]
    hdr_t = next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True))
    ti = {(h or "").strip(): i for i, h in enumerate(hdr_t)}
    trade_dir: dict = {}
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        tid = row[ti["Trade ID"]]
        if not tid:
            continue
        sym_full = row[ti["Instrument Symbol"]] or ""
        trade_dir[tid] = {
            "oc": (row[ti["Open/Close"]] or "").strip(),
            "sym": sym_full.split(":")[0],
        }

    # ── Bookings: aggregate "Commission" rows by ticker × direction ──
    ws_b = wb["Bookings"]
    hdr_b = next(ws_b.iter_rows(min_row=1, max_row=1, values_only=True))
    bi = {(h or "").strip(): i for i, h in enumerate(hdr_b)}

    open_comm: dict[str, float] = defaultdict(float)
    close_comm: dict[str, float] = defaultdict(float)
    unmatched = 0.0
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        at = (row[bi["Amount Type"]] or "").strip()
        if at != "Commission":
            continue
        tid = row[bi["Trade ID"]]
        amt = row[bi["Booked Amount"]] or 0
        info = trade_dir.get(tid)
        if not info:
            unmatched += amt
            continue
        tk = info["sym"]
        if info["oc"] == "To Close":
            close_comm[tk] += amt
        else:  # "To Open" or anything else → treat as open-side
            open_comm[tk] += amt

    if unmatched:
        print(
            f"  WARN  ${unmatched:.2f} of commission could not be matched to a Trade "
            "(probably non-Trade booking rows). Skipped.",
            file=sys.stderr,
        )

    return dict(open_comm), dict(close_comm)


def get_close_tickers(xl_path: Path) -> set[str]:
    """
    Returns set of ticker symbols that have at least one 'To Close' trade in xlsx.
    Used to identify phantom closed records created by the PDF parser for open-only positions.
    """
    wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
    ws_t = wb["Trades"]
    hdr_t = next(ws_t.iter_rows(min_row=1, max_row=1, values_only=True))
    ti = {(h or "").strip(): i for i, h in enumerate(hdr_t)}
    close_tickers: set[str] = set()
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        oc = (row[ti["Open/Close"]] or "").strip()
        sym_full = row[ti["Instrument Symbol"]] or ""
        if oc == "To Close":
            close_tickers.add(sym_full.split(":")[0])
    return close_tickers


def patch_holdings(
    cost: dict, open_comm: dict[str, float], close_comm: dict[str, float],
    close_tickers: set[str] | None = None,
) -> tuple[dict, dict]:
    """Apply commission overlay + remove phantom closed records. Returns (new_cost, diff_summary)."""
    us = cost.setdefault("us", {})
    diff: dict = {"open": {}, "closed": {}, "phantom_removed": []}

    # ── us.open[] fees overlay ──
    for p in us.get("open", []):
        tk = p.get("tk")
        if tk and tk in open_comm:
            new = round(open_comm[tk], 2)
            old = p.get("fees", 0.0)
            if abs(new - old) > 0.005:
                diff["open"][tk] = {"old": old, "new": new}
            p["fees"] = new

    # ── Remove phantom closed records ──
    # PDF parser creates phantom closed records for open positions when P/L parsing
    # yields a non-zero realised_est. Rule: if ticker is currently OPEN and has
    # NO "To Close" trade in xlsx → never actually sold → remove phantom.
    if close_tickers is not None:
        open_tks = {h.get("tk") for h in us.get("open", []) if h.get("tk")}
        keep, phantoms = [], []
        for p in us.get("closed", []):
            tk = p.get("tk", "")
            is_phantom = (
                tk in open_tks                  # currently open
                and tk not in close_tickers     # never sold in xlsx window
                and not tk.startswith("UNKNOWN")
            )
            if is_phantom:
                phantoms.append(tk)
                print(f"  PHANTOM REMOVED {tk}: open position, no sell in xlsx")
            else:
                keep.append(p)
        us["closed"] = keep
        diff["phantom_removed"] = phantoms

    # ── us.closed[] _costs_paid overlay ──
    # Only patch entries whose ticker has trades in the xlsx window.
    # Realised P&L is already net of fees in the PDF — do NOT touch realised.
    #
    # Two cases:
    #   (a) ticker is in BOTH us.open and us.closed (re-opened) →
    #       open lot fees = open_comm,  closed _costs_paid = close_comm.
    #   (b) ticker is in us.closed ONLY (fully closed in xlsx window) →
    #       _costs_paid = open_comm + close_comm  (entire ticker's fees belong
    #       to the closed tranche, since no open lot exists to receive them).
    open_tks = {h.get("tk") for h in us.get("open", []) if h.get("tk")}
    for p in us.get("closed", []):
        tk = p.get("tk")
        if not tk:
            continue
        if tk in open_tks:
            # Re-opened ticker → only close-side commission belongs to the closed tranche
            new_costs = round(close_comm.get(tk, 0.0), 2)
            in_xlsx = tk in close_comm
        else:
            # Closed-only ticker → both open- and close-side commissions belong here
            new_costs = round(open_comm.get(tk, 0.0) + close_comm.get(tk, 0.0), 2)
            in_xlsx = (tk in open_comm) or (tk in close_comm)
        if in_xlsx:
            old = p.get("_costs_paid", 0.0)
            if abs(new_costs - old) > 0.005:
                diff["closed"][tk] = {"old": old, "new": new_costs}
            p["_costs_paid"] = new_costs

    return cost, diff


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Path to US broker transactions xlsx")
    ap.add_argument(
        "--cost-file",
        default=str(ROOT / "data" / "holdings_cost.json"),
        help="Path to holdings_cost.json (default: data/holdings_cost.json)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Print diff but do not write the file",
    )
    args = ap.parse_args()

    xl = Path(args.xlsx).expanduser()
    cost_path = Path(args.cost_file).expanduser()
    if not xl.exists():
        sys.exit(f"xlsx not found: {xl}")
    if not cost_path.exists():
        sys.exit(f"cost file not found: {cost_path}")

    print(f"Reading commissions from {xl.name}…")
    open_comm, close_comm = split_commissions(xl)
    close_tickers = get_close_tickers(xl)
    print(
        f"  {len(open_comm)} tickers with open-side commission, "
        f"{len(close_comm)} with close-side"
    )
    print(f"  sum Open:   ${sum(open_comm.values()):,.2f}")
    print(f"  sum Close:  ${sum(close_comm.values()):,.2f}")
    print(f"  grand tot:  ${sum(open_comm.values()) + sum(close_comm.values()):,.2f}")
    print(f"  tickers with sell trades: {sorted(close_tickers)}")

    cost = json.loads(cost_path.read_text())
    patched, diff = patch_holdings(cost, open_comm, close_comm, close_tickers=close_tickers)

    if diff["open"] or diff["closed"]:
        print("\n── Changes ──")
        for tk, d in sorted(diff["open"].items()):
            print(f"  open    {tk:8s}  fees       ${d['old']:>8,.2f}  →  ${d['new']:>8,.2f}")
        for tk, d in sorted(diff["closed"].items()):
            print(f"  closed  {tk:8s}  _costs_paid ${d['old']:>8,.2f}  →  ${d['new']:>8,.2f}")
    else:
        print("\nNo changes — fees already match xlsx.")

    if args.dry_run:
        print("\n(dry-run — file NOT written)")
        return 0

    cost_path.write_text(json.dumps(patched, indent=2, ensure_ascii=False))
    print(f"\nwrote {cost_path.relative_to(ROOT) if cost_path.is_relative_to(ROOT) else cost_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
