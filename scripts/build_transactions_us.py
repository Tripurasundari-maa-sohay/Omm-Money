"""
build_transactions_us.py — produce data/transactions_us.json from the Saxo/Doha
transactions xlsx.

Each output row is a single trade with full attribution: date, ticker, side,
qty, price, gross, commission, FTT, net cash impact, open/close direction,
realised P/L (for sells), name, ISIN, exchange, asset type. Cash movements
(deposits/withdrawals, dividends, interest, custody fees) are listed
separately under cash_movements[].

The output feeds the dashboard Transactions tab "Per-trade ledger" and is
also intended as ML/training input later (every trade with deterministic
schema, no PII).

Usage:
    python3 scripts/build_transactions_us.py <path-to-transactions.xlsx>
        [--out data/transactions_us.json]
        [--dry-run]
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = ROOT / "data" / "transactions_us.json"


def _iso_date(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    try:
        return datetime.fromisoformat(str(v).split(" ")[0]).date().isoformat()
    except Exception:
        return str(v)


def _base_symbol(s: str | None) -> str:
    return s.split(":")[0] if s else ""


def _header_index(ws) -> dict[str, int]:
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {(h or "").strip(): i for i, h in enumerate(hdr)}


def build(xl_path: Path) -> dict:
    wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
    for need in ("Transactions", "Trades", "Bookings"):
        if need not in wb.sheetnames:
            raise SystemExit(f"xlsx missing required sheet: {need!r} (found {wb.sheetnames})")

    # ── Trades sheet: open/close + traded value reference ──
    ws_t = wb["Trades"]
    ti = _header_index(ws_t)
    trade_meta: dict = {}
    for row in ws_t.iter_rows(min_row=2, values_only=True):
        tid = row[ti["Trade ID"]]
        if not tid:
            continue
        oc_raw = (row[ti["Open/Close"]] or "").strip().lower()
        oc = "open" if "open" in oc_raw else ("close" if "close" in oc_raw else None)
        trade_meta[tid] = {
            "oc": oc,
            "exch": row[ti["Exchange Description"]] or "",
            "order_type": row[ti["Trade Type"]] or "",
        }

    # ── Bookings sheet: per-trade fee components (commission, FTT, etc.) ──
    ws_b = wb["Bookings"]
    bi = _header_index(ws_b)
    fees_by_trade: dict = defaultdict(lambda: defaultdict(float))
    # Cash movements not attached to a Trade ID — sourced from Bookings since
    # the Transactions sheet's "Corporate action" rows lack the granular type
    # (Dividend vs Withholding vs Custody).
    standalone_cash = []
    for row in ws_b.iter_rows(min_row=2, values_only=True):
        at = (row[bi["Amount Type"]] or "").strip()
        tid = row[bi["Trade ID"]]
        amt = row[bi["Booked Amount"]] or 0
        date = _iso_date(row[bi["Booking date"]])
        if tid:
            if at == "Commission":
                fees_by_trade[tid]["commission"] += amt
            elif at == "French Financial Transaction Tax":
                fees_by_trade[tid]["ftt"] += amt
            elif at == "Exchange Fee":
                fees_by_trade[tid]["exchange_fee"] += amt
            elif at == "Share Amount":
                fees_by_trade[tid]["share_amt"] += amt
            elif at == "Corporate Actions - Withholding Tax":
                fees_by_trade[tid]["withholding"] += amt
            # else: ignore non-trade booking types here
        else:
            # No Trade ID → standalone cash movement (custody fee, dividend, interest, withholding)
            # Skip "Share Amount" + "Cash Amount" — these are double-entries that
            # mirror the canonical Cash Transfer row already captured from the
            # Transactions sheet.
            if at and at not in ("Share Amount", "Cash Amount"):
                standalone_cash.append({
                    "date": date,
                    "type": at,
                    "amount": round(amt, 4),
                    "instrument": row[bi["Instrument"]] or None,
                    "symbol": _base_symbol(row[bi["Instrument Symbol"]]) or None,
                })

    # ── Transactions sheet: one row per trade or cash transfer (canonical) ──
    ws_x = wb["Transactions"]
    xi = _header_index(ws_x)
    trades: list[dict] = []
    cash_moves: list[dict] = []
    for row in ws_x.iter_rows(min_row=2, values_only=True):
        ttype = (row[xi["Transaction Type"]] or "").strip()
        date  = _iso_date(row[xi["Trade Date"]])
        event = (row[xi["Event"]] or "").strip()
        booked = row[xi["Booked Amount"]] or 0
        ccy   = row[xi["Currency"]] or "USD"
        sym   = _base_symbol(row[xi["Instrument Symbol"]])
        name  = (row[xi["Instrument"]] or "").strip()
        isin  = row[xi["Instrument ISIN"]] or None
        atype = row[xi["Type"]] or None
        exch  = row[xi["Exchange Description"]] or None
        rpnl  = row[xi["Realized P/L"]]
        tid   = row[xi["Trade ID"]]

        if ttype == "Trade" and tid:
            # Parse "Buy 25 @ 38.10 USD" / "Sell -11 @ 178.62 USD"
            side = "buy" if event.lower().startswith("buy") else "sell" if event.lower().startswith("sell") else None
            qty = price = None
            try:
                parts = event.split()
                # parts ~ ['Buy', '25', '@', '38.10', 'USD']  or  ['Sell', '-11', '@', '178.62', 'USD']
                qty = abs(float(parts[1]))
                price = float(parts[3])
            except Exception:
                pass
            fees = fees_by_trade.get(tid, {})
            commission   = round(fees.get("commission", 0.0), 4)
            ftt          = round(fees.get("ftt", 0.0), 4)
            exchange_fee = round(fees.get("exchange_fee", 0.0), 4)
            gross = round(qty * price, 4) if (qty is not None and price is not None) else None
            net   = round(booked, 4)  # booked amount = cash impact (includes commission)
            meta  = trade_meta.get(tid, {})
            trades.append({
                "trade_id":     str(tid),
                "date":         date,
                "tk":           sym,
                "side":         side,
                "qty":          qty,
                "price":        price,
                "gross":        gross,
                "commission":   commission,
                "ftt":          ftt,
                "exchange_fee": exchange_fee,
                "net":          net,
                "open_close":   meta.get("oc"),
                "order_type":   meta.get("order_type") or None,
                "realised_pl":  round(rpnl, 4) if isinstance(rpnl, (int, float)) else None,
                "name":         name,
                "isin":         isin,
                "ccy":          ccy,
                "asset_type":   atype,
                "exchange":     exch or meta.get("exch") or None,
            })
        else:
            # Non-Trade Transaction rows: keep only "Cash Transfer" (deposits/
            # withdrawals). Skip:
            #   • "Cash Amount" / "Cash amount" — Share-Amount counterpart, duplicate
            #   • "Corporate action" — dividends are captured from Bookings
            #     (standalone_cash) with finer-grained type (Cash Dividend vs
            #     Withholding) so we avoid double-listing here.
            if ttype != "Cash Transfer":
                continue
            cash_moves.append({
                "date":   date,
                "type":   ttype,
                "event":  event or None,
                "amount": round(booked, 4),
                "ccy":    ccy,
                "tk":     sym or None,
                "name":   name or None,
            })

    # Merge standalone Booking rows (custody fee, dividends without Trade ID) into cash_moves
    for c in standalone_cash:
        cash_moves.append({
            "date":   c["date"],
            "type":   c["type"],
            "event":  None,
            "amount": c["amount"],
            "ccy":    "USD",
            "tk":     c["symbol"],
            "name":   c["instrument"],
        })

    # ── Sort newest-first ──
    trades.sort(key=lambda r: (r["date"] or "", r["trade_id"]), reverse=True)
    cash_moves.sort(key=lambda r: (r["date"] or ""), reverse=True)

    # ── Totals ──
    buys  = [t for t in trades if t["side"] == "buy"]
    sells = [t for t in trades if t["side"] == "sell"]
    gross_buys  = round(sum(t["gross"] or 0 for t in buys), 2)
    gross_sells = round(sum(t["gross"] or 0 for t in sells), 2)
    commission_total = round(sum(t["commission"] for t in trades), 2)
    realised_total   = round(sum(t["realised_pl"] or 0 for t in sells), 2)
    deposits = round(sum(c["amount"] for c in cash_moves if c["type"] == "Cash Transfer" and c["amount"] > 0), 2)
    withdrawals = round(sum(c["amount"] for c in cash_moves if c["type"] == "Cash Transfer" and c["amount"] < 0), 2)
    dividends = round(sum(c["amount"] for c in cash_moves if "Dividend" in (c["type"] or "")), 2)
    interest  = round(sum(c["amount"] for c in cash_moves if "Interest" in (c["type"] or "")), 2)
    custody   = round(sum(c["amount"] for c in cash_moves if "Custody" in (c["type"] or "")), 2)
    withhold  = round(sum(c["amount"] for c in cash_moves if "Withholding" in (c["type"] or "")), 2)

    period_start = min((t["date"] for t in trades if t["date"]), default=None)
    period_end   = max((t["date"] for t in trades if t["date"]), default=None)

    return {
        "generated":  datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "source":     "Saxo / Doha Bank transactions xlsx",
        "period_from": period_start,
        "period_to":   period_end,
        "trades":      trades,
        "cash_moves":  cash_moves,
        "totals": {
            "trades_count":      len(trades),
            "buys_count":        len(buys),
            "sells_count":       len(sells),
            "gross_buys":        gross_buys,
            "gross_sells":       gross_sells,
            "commission_total":  commission_total,
            "realised_pl_total": realised_total,
            "deposits":          deposits,
            "withdrawals":       withdrawals,
            "dividends":         dividends,
            "interest":          interest,
            "custody_fees":      custody,
            "withholding_tax":   withhold,
        },
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Saxo/Doha transactions xlsx")
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    xl = Path(args.xlsx).expanduser()
    if not xl.exists():
        sys.exit(f"xlsx not found: {xl}")

    print(f"Building transactions from {xl.name}…")
    data = build(xl)
    t = data["totals"]
    print(
        f"  trades:      {t['trades_count']:>4d}  "
        f"(buys {t['buys_count']}, sells {t['sells_count']})\n"
        f"  gross buys:  ${t['gross_buys']:>11,.2f}\n"
        f"  gross sells: ${t['gross_sells']:>11,.2f}\n"
        f"  commission:  ${t['commission_total']:>11,.2f}\n"
        f"  realised:    ${t['realised_pl_total']:>11,.2f}\n"
        f"  deposits:    ${t['deposits']:>11,.2f}\n"
        f"  dividends:   ${t['dividends']:>11,.2f}\n"
        f"  custody:     ${t['custody_fees']:>11,.2f}\n"
        f"  period:      {data['period_from']} → {data['period_to']}"
    )

    if args.dry_run:
        print("(dry-run — not written)")
        return 0

    out_path = Path(args.out).expanduser()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f"\nwrote {out_path.relative_to(ROOT) if out_path.is_relative_to(ROOT) else out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
